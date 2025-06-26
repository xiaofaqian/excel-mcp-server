"""
Insert Cell Text Tool - Excel单元格文本插入工具
提供向指定单元格插入文本的功能，支持精确定位和格式控制
"""

import logging
import os
from typing import Annotated, Optional, Dict, Any
from pydantic import Field
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# 配置日志
logger = logging.getLogger("excel-mcp-server")


def _validate_parameters(file_path: str, row_number: int, column_number: int, text_content: str) -> Optional[str]:
    """验证输入参数"""
    # 验证文件路径
    if not os.path.exists(file_path):
        return f"文件不存在: {file_path}"
    
    # 验证文件扩展名
    file_ext = os.path.splitext(file_path)[1].lower()
    if file_ext not in ['.xlsx', '.xls']:
        return f"不支持的文件格式: {file_ext}。仅支持 .xlsx 和 .xls 文件"
    
    # 验证行号和列号
    if row_number < 1:
        return f"行号必须大于等于1，当前值: {row_number}"
    
    if column_number < 1:
        return f"列号必须大于等于1，当前值: {column_number}"
    
    if row_number > 1048576:  # Excel最大行数
        return f"行号不能超过1048576，当前值: {row_number}"
    
    if column_number > 16384:  # Excel最大列数
        return f"列号不能超过16384，当前值: {column_number}"
    
    # 验证文本内容
    if not isinstance(text_content, str):
        return f"文本内容必须是字符串类型，当前类型: {type(text_content)}"
    
    # 防止公式注入
    if text_content.strip().startswith('='):
        return "不允许插入公式，此工具仅支持纯文本插入"
    
    # 文本长度限制
    if len(text_content) > 32767:  # Excel单元格文本长度限制
        return f"文本长度不能超过32767个字符，当前长度: {len(text_content)}"
    
    return None


def _apply_cell_formatting(cell, preserve_formatting: bool, original_cell=None):
    """应用单元格格式，强制左对齐"""
    try:
        # 强制设置左对齐（这是主要需求）
        cell.alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=False
        )
        
        # 如果需要保持格式且有原始单元格，则复制其他格式但保持左对齐
        if preserve_formatting and original_cell:
            # 复制字体格式
            if original_cell.font:
                cell.font = original_cell.font
            
            # 复制填充格式
            if original_cell.fill:
                cell.fill = original_cell.fill
            
            # 复制边框格式
            if original_cell.border:
                cell.border = original_cell.border
            
            # 复制数字格式
            if original_cell.number_format:
                cell.number_format = original_cell.number_format
            
            # 注意：不复制对齐方式，因为我们要强制使用左对齐
        
    except Exception as e:
        logger.warning(f"[Warning] 应用单元格格式时出现警告: {str(e)}")


def insert_cell_text(
    file_path: Annotated[str, Field(
        description="Excel 文件的完整路径（支持 .xlsx 和 .xls 格式）"
    )],
    row_number: Annotated[int, Field(
        description="目标单元格的行号，从1开始计数"
    )],
    column_number: Annotated[int, Field(
        description="目标单元格的列号，从1开始计数"
    )],
    text_content: Annotated[str, Field(
        description="要插入的文本内容，仅支持纯文本，不支持公式"
    )],
    sheet_name: Annotated[Optional[str], Field(
        default=None,
        description="要操作的工作表名称。如果不指定，将操作第一个工作表"
    )] = None,
    preserve_formatting: Annotated[bool, Field(
        default=True,
        description="是否保持原有的单元格格式（字体、颜色、边框等），对齐方式始终设置为左对齐"
    )] = True
) -> Dict[str, Any]:
    """向指定的单元格插入文本，默认左对齐"""
    
    logger.info(f"[Tool] insert_cell_text called with file_path: {file_path}, row: {row_number}, col: {column_number}")
    
    try:
        # 1. 参数验证
        param_error = _validate_parameters(file_path, row_number, column_number, text_content)
        if param_error:
            logger.error(f"[Error] {param_error}")
            return {
                "success": False,
                "error": param_error,
                "data": None
            }
        
        logger.info(f"[API] 准备向单元格 ({row_number}, {column_number}) 插入文本: '{text_content[:50]}{'...' if len(text_content) > 50 else ''}'")
        
        # 2. 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        
        # 确定目标工作表
        if sheet_name and sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            target_sheet = sheet_name
        elif sheet_name:
            error_msg = f"工作表 '{sheet_name}' 不存在。可用工作表: {workbook.sheetnames}"
            logger.error(f"[Error] {error_msg}")
            return {
                "success": False,
                "error": error_msg,
                "data": None
            }
        else:
            worksheet = workbook.active
            target_sheet = worksheet.title
        
        logger.info(f"[API] 操作工作表: {target_sheet}")
        
        # 3. 获取目标单元格
        target_cell = worksheet.cell(row=row_number, column=column_number)
        
        # 记录原始值
        original_value = target_cell.value
        
        # 获取单元格位置标识（如A1, B2等）
        cell_position = f"{get_column_letter(column_number)}{row_number}"
        
        # 4. 保存原始格式信息（如果需要保持格式）
        original_cell_copy = None
        if preserve_formatting:
            # 创建一个临时单元格来保存原始格式
            temp_cell = worksheet.cell(row=row_number, column=column_number)
            original_cell_copy = temp_cell
        
        # 5. 插入文本内容
        target_cell.value = text_content
        logger.info(f"[API] 文本已插入到单元格 {cell_position}")
        
        # 6. 应用格式（强制左对齐）
        _apply_cell_formatting(target_cell, preserve_formatting, original_cell_copy)
        logger.info(f"[API] 格式已应用到单元格 {cell_position}，左对齐已设置")
        
        # 7. 保存文件
        workbook.save(file_path)
        logger.info(f"[API] 文件已保存: {file_path}")
        
        # 8. 构建返回结果
        result = {
            "success": True,
            "error": None,
            "data": {
                "file_path": file_path,
                "sheet_name": target_sheet,
                "cell_position": cell_position,
                "row_number": row_number,
                "column_number": column_number,
                "text_content": text_content,
                "text_length": len(text_content),
                "left_aligned": True,
                "formatting_preserved": preserve_formatting,
                "original_value": original_value,
                "cell_existed": original_value is not None
            }
        }
        
        logger.info(f"[Tool] 成功向单元格 {cell_position} 插入文本，长度: {len(text_content)} 字符")
        return result
        
    except FileNotFoundError:
        error_msg = f"文件未找到: {file_path}"
        logger.error(f"[Error] {error_msg}")
        return {
            "success": False,
            "error": error_msg,
            "data": None
        }
    except PermissionError:
        error_msg = f"没有权限访问文件: {file_path}"
        logger.error(f"[Error] {error_msg}")
        return {
            "success": False,
            "error": error_msg,
            "data": None
        }
    except Exception as e:
        error_msg = f"插入单元格文本时发生未知错误: {str(e)}"
        logger.error(f"[Error] {error_msg}")
        return {
            "success": False,
            "error": error_msg,
            "data": None
        }
