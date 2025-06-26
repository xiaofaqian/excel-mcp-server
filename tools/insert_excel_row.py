"""
Insert Excel Row Tool - Excel行插入工具
提供Excel文件中插入行数据的功能，支持批量插入、数据验证、公式计算和格式保持
优化版本：支持左对齐、空行清理、代码结构重构
"""

import logging
import os
import re
from typing import Annotated, Optional, Dict, Any, List, Union, Tuple
from pydantic import Field
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import time
import tracemalloc

# 配置日志
logger = logging.getLogger("excel-mcp-server")


def _validate_parameters(file_path: str, batch_size: int) -> Optional[str]:
    """验证输入参数"""
    # 验证 batch_size 参数
    if batch_size > 500:
        return f"batch_size 参数不能超过 500，当前值: {batch_size}"
    
    if batch_size <= 0:
        return f"batch_size 参数必须大于 0，当前值: {batch_size}"
    
    # 验证文件路径
    if not os.path.exists(file_path):
        return f"文件不存在: {file_path}"
    
    # 验证文件扩展名
    file_ext = os.path.splitext(file_path)[1].lower()
    if file_ext not in ['.xlsx', '.xls']:
        return f"不支持的文件格式: {file_ext}。仅支持 .xlsx 和 .xls 文件"
    
    return None


def _validate_data(rows_to_insert: List[Dict[str, Any]], validation_rules: Optional[Dict[str, Dict[str, Any]]]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """验证数据并返回验证通过的行和验证报告"""
    validation_report = {"passed": 0, "failed": 0, "errors": []}
    
    if not validation_rules:
        validation_report["passed"] = len(rows_to_insert)
        return rows_to_insert, validation_report
    
    logger.info("[API] Performing data validation")
    validated_rows = []
    
    for row_idx, row in enumerate(rows_to_insert):
        row_errors = []
        validated_row = row.copy()
        
        # 验证每个字段
        for col_name, rules in validation_rules.items():
            if col_name in row:
                value = row[col_name]
                
                # 必填验证
                if rules.get('required', False) and (value is None or str(value).strip() == ''):
                    row_errors.append(f"列 '{col_name}' 为必填项")
                    continue
                
                if value is not None and str(value).strip() != '':
                    # 类型验证
                    data_type = rules.get('type', 'string')
                    if data_type == 'number':
                        try:
                            num_value = float(value)
                            if 'min_value' in rules and num_value < rules['min_value']:
                                row_errors.append(f"列 '{col_name}' 值 {num_value} 小于最小值 {rules['min_value']}")
                            if 'max_value' in rules and num_value > rules['max_value']:
                                row_errors.append(f"列 '{col_name}' 值 {num_value} 大于最大值 {rules['max_value']}")
                        except ValueError:
                            row_errors.append(f"列 '{col_name}' 不是有效的数字")
                    
                    elif data_type == 'email':
                        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                        if not re.match(email_pattern, str(value)):
                            row_errors.append(f"列 '{col_name}' 不是有效的邮箱格式")
                    
                    # 自定义正则表达式验证
                    if 'pattern' in rules:
                        if not re.match(rules['pattern'], str(value)):
                            row_errors.append(f"列 '{col_name}' 不符合指定格式")
                    
                    # 字符串长度验证
                    if data_type == 'string':
                        str_value = str(value)
                        if 'min_length' in rules and len(str_value) < rules['min_length']:
                            row_errors.append(f"列 '{col_name}' 长度不能少于 {rules['min_length']} 个字符")
                        if 'max_length' in rules and len(str_value) > rules['max_length']:
                            row_errors.append(f"列 '{col_name}' 长度不能超过 {rules['max_length']} 个字符")
        
        if row_errors:
            validation_report["failed"] += 1
            validation_report["errors"].extend([f"第{row_idx+1}行: {error}" for error in row_errors])
        else:
            validation_report["passed"] += 1
            validated_rows.append(validated_row)
    
    logger.info(f"[API] Validation completed. Passed: {validation_report['passed']}, Failed: {validation_report['failed']}")
    return validated_rows, validation_report


def _determine_insert_position(insert_position: str, max_row: int) -> Tuple[int, Optional[str]]:
    """确定插入位置"""
    if insert_position == "end":
        return max_row + 1, None
    elif insert_position == "beginning":
        return 2, None  # 在标题行后插入
    elif insert_position.startswith("after_row_"):
        try:
            target_row = int(insert_position.split("_")[-1])
            return target_row + 1, None
        except ValueError:
            return 0, f"无效的插入位置格式: {insert_position}"
    else:
        return 0, f"不支持的插入位置: {insert_position}"


def _apply_cell_alignment(cell, preserve_formatting: bool, reference_cell=None):
    """应用单元格对齐设置，优先设置左对齐"""
    try:
        # 设置左对齐（这是主要需求）
        cell.alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=False
        )
        
        # 如果需要保持格式且有参考单元格，则复制其他格式但保持左对齐
        if preserve_formatting and reference_cell:
            if reference_cell.font:
                cell.font = Font(
                    name=reference_cell.font.name,
                    size=reference_cell.font.size,
                    bold=reference_cell.font.bold,
                    italic=reference_cell.font.italic,
                    color=reference_cell.font.color
                )
            if reference_cell.fill:
                cell.fill = PatternFill(
                    fill_type=reference_cell.fill.fill_type,
                    start_color=reference_cell.fill.start_color,
                    end_color=reference_cell.fill.end_color
                )
            if reference_cell.border:
                cell.border = Border(
                    left=reference_cell.border.left,
                    right=reference_cell.border.right,
                    top=reference_cell.border.top,
                    bottom=reference_cell.border.bottom
                )
            # 注意：这里不复制对齐方式，因为我们要强制使用左对齐
    except Exception as e:
        logger.warning(f"[Warning] Could not apply formatting to cell: {str(e)}")


def _detect_empty_rows(worksheet) -> List[int]:
    """检测工作表中的空行"""
    empty_rows = []
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    # 从第2行开始检查（跳过标题行）
    for row_num in range(2, max_row + 1):
        is_empty = True
        for col_num in range(1, max_col + 1):
            cell_value = worksheet.cell(row_num, col_num).value
            if cell_value is not None and str(cell_value).strip():
                is_empty = False
                break
        
        if is_empty:
            empty_rows.append(row_num)
    
    return empty_rows


def _remove_empty_rows(worksheet, empty_rows: List[int]) -> int:
    """删除空行，返回删除的行数"""
    if not empty_rows:
        return 0
    
    # 从后往前删除，避免行号变化
    empty_rows.sort(reverse=True)
    removed_count = 0
    
    for row_num in empty_rows:
        worksheet.delete_rows(row_num)
        removed_count += 1
    
    return removed_count


def _insert_data_with_formatting(worksheet, rows_to_insert: List[Dict[str, Any]], headers: List[str], 
                                insert_row: int, preserve_formatting: bool, calculate_formulas: bool, 
                                reference_row: Optional[int]) -> Dict[str, Any]:
    """插入数据并应用格式"""
    formula_report = {"processed": 0, "adjusted": 0}
    
    for row_idx, row_data_dict in enumerate(rows_to_insert):
        current_row = insert_row + row_idx
        
        # 为每一列设置数据
        for col_idx, header in enumerate(headers):
            col_letter = get_column_letter(col_idx + 1)
            cell = worksheet[f"{col_letter}{current_row}"]
            
            # 设置数据值
            if header in row_data_dict:
                value = row_data_dict[header]
                
                # 处理公式
                if calculate_formulas and isinstance(value, str) and value.startswith('='):
                    formula_report["processed"] += 1
                    # 调整公式中的行引用
                    adjusted_formula = _adjust_formula_references(value, current_row, worksheet.max_row)
                    if adjusted_formula != value:
                        formula_report["adjusted"] += 1
                    cell.value = adjusted_formula
                    logger.debug(f"[API] Set formula in {col_letter}{current_row}: {adjusted_formula}")
                else:
                    cell.value = value
            else:
                cell.value = None
            
            # 应用格式和对齐
            reference_cell = worksheet[f"{col_letter}{reference_row}"] if reference_row else None
            _apply_cell_alignment(cell, preserve_formatting, reference_cell)
    
    return formula_report


def _adjust_formula_references(formula: str, current_row: int, original_max_row: int) -> str:
    """
    调整公式中的行引用
    将相对引用调整为适合新插入行的位置
    """
    try:
        # 简单的行引用调整逻辑
        # 这里实现基本的行号调整，更复杂的公式可能需要更高级的解析
        
        # 匹配形如 A1, B2, $A$1 等单元格引用
        cell_pattern = r'([A-Z]+)(\d+)'
        
        def replace_cell_ref(match):
            col = match.group(1)
            row = int(match.group(2))
            
            # 如果引用的行在插入位置之后，需要调整行号
            if row > original_max_row:
                # 这是一个相对引用，需要调整
                adjusted_row = row + (current_row - original_max_row - 1)
                return f"{col}{adjusted_row}"
            else:
                # 保持原有引用
                return match.group(0)
        
        adjusted_formula = re.sub(cell_pattern, replace_cell_ref, formula)
        return adjusted_formula
        
    except Exception as e:
        logger.warning(f"[Warning] Could not adjust formula '{formula}': {str(e)}")
        return formula


def insert_excel_row(
    file_path: Annotated[str, Field(
        description="Excel 文件的完整路径（支持 .xlsx 和 .xls 格式）"
    )],
    row_data: Annotated[Union[Dict[str, Any], List[Dict[str, Any]]], Field(
        description="要插入的行数据。可以是单行数据（字典）或多行数据（字典列表）"
    )],
    sheet_name: Annotated[Optional[str], Field(
        default=None,
        description="要操作的工作表名称。如果不指定，将操作第一个工作表"
    )] = None,
    insert_position: Annotated[str, Field(
        default="end",
        description="插入位置：'end'(末尾)、'beginning'(开头) 或 'after_row_N'(在第N行后)"
    )] = "end",
    validation_rules: Annotated[Optional[Dict[str, Dict[str, Any]]], Field(
        default=None,
        description="数据验证规则，格式：{'列名': {'type': 'string|number|date|email', 'required': bool, 'min_value': float, 'max_value': float, 'pattern': 'regex'}}"
    )] = None,
    preserve_formatting: Annotated[bool, Field(
        default=True,
        description="是否保持原有的单元格格式（字体、颜色、边框等）"
    )] = True,
    calculate_formulas: Annotated[bool, Field(
        default=True,
        description="是否处理和调整公式中的行引用"
    )] = True,
    batch_size: Annotated[int, Field(
        default=100,
        description="批量处理的最大行数，防止内存溢出。默认100行，最多不能超过500行"
    )] = 100,
    save_as: Annotated[Optional[str], Field(
        default=None,
        description="另存为新文件的路径。如果不指定，将覆盖原文件"
    )] = None
) -> Dict[str, Any]:
    """在Excel文件中插入一行或多行数据，支持批量插入、数据验证、公式计算、格式保持、左对齐和空行清理"""
    
    # 开始性能监控
    tracemalloc.start()
    start_time = time.time()
    
    logger.info(f"[Tool] insert_excel_row called with file_path: {file_path}, rows: {len(row_data) if isinstance(row_data, list) else 1}")
    
    try:
        # 1. 参数验证
        param_error = _validate_parameters(file_path, batch_size)
        if param_error:
            logger.error(f"[Error] {param_error}")
            return {
                "success": False,
                "error": param_error,
                "data": None
            }
        
        # 2. 标准化行数据格式
        if isinstance(row_data, dict):
            rows_to_insert = [row_data]
        else:
            rows_to_insert = row_data
        
        # 检查批量大小限制
        if len(rows_to_insert) > batch_size:
            error_msg = f"要插入的行数 ({len(rows_to_insert)}) 超过批量处理限制 ({batch_size})"
            logger.error(f"[Error] {error_msg}")
            return {
                "success": False,
                "error": error_msg,
                "data": None
            }
        
        logger.info(f"[API] Processing {len(rows_to_insert)} rows for insertion")
        
        # 3. 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        
        # 确定目标工作表
        if sheet_name and sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            target_sheet = sheet_name
        elif sheet_name:
            error_msg = f"工作表 '{sheet_name}' 不存在。可用工作表: {workbook.sheetnames}"
            logger.error(f"[Error] {error_msg}")
            return {
                "success": False,
                "error": error_msg,
                "data": None
            }
        else:
            worksheet = workbook.active
            target_sheet = worksheet.title
        
        logger.info(f"[API] Working with sheet: {target_sheet}")
        
        # 4. 获取现有数据和列标题
        data_rows = list(worksheet.iter_rows(values_only=True))
        if not data_rows:
            error_msg = "工作表为空，无法确定列结构"
            logger.error(f"[Error] {error_msg}")
            return {
                "success": False,
                "error": error_msg,
                "data": None
            }
        
        # 获取列标题（第一行）
        headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(data_rows[0])]
        logger.info(f"[API] Found headers: {headers}")
        
        # 5. 数据验证
        validated_rows, validation_report = _validate_data(rows_to_insert, validation_rules)
        
        if not validated_rows:
            error_msg = "所有行都未通过验证，没有数据可插入"
            logger.error(f"[Error] {error_msg}")
            return {
                "success": False,
                "error": error_msg,
                "data": validation_report
            }
        
        # 6. 确定插入位置
        max_row = worksheet.max_row
        insert_row, position_error = _determine_insert_position(insert_position, max_row)
        if position_error:
            logger.error(f"[Error] {position_error}")
            return {
                "success": False,
                "error": position_error,
                "data": None
            }
        
        logger.info(f"[API] Inserting rows starting at row {insert_row}")
        
        # 7. 如果不是在末尾插入，需要先插入空行
        if insert_position != "end":
            for i in range(len(validated_rows)):
                worksheet.insert_rows(insert_row + i)
        
        # 8. 获取参考行的格式（用于格式保持）
        reference_row = max(2, min(max_row, insert_row - 1)) if preserve_formatting else None
        
        # 9. 插入数据并应用格式
        formula_report = _insert_data_with_formatting(
            worksheet, validated_rows, headers, insert_row, 
            preserve_formatting, calculate_formulas, reference_row
        )
        
        # 10. 检测和清理空行
        logger.info("[API] Detecting and removing empty rows")
        empty_rows = _detect_empty_rows(worksheet)
        removed_rows = _remove_empty_rows(worksheet, empty_rows)
        
        if removed_rows > 0:
            logger.info(f"[API] Removed {removed_rows} empty rows")
        
        # 11. 保存文件
        output_file = save_as if save_as else file_path
        workbook.save(output_file)
        logger.info(f"[API] File saved to: {output_file}")
        
        # 12. 性能统计
        end_time = time.time()
        current, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        
        processing_time = round(end_time - start_time, 2)
        memory_usage = round(peak / 1024 / 1024, 2)  # MB
        
        # 13. 构建返回结果
        result = {
            "success": True,
            "error": None,
            "data": {
                "file_path": file_path,
                "output_file": output_file,
                "sheet_name": target_sheet,
                "inserted_rows": len(validated_rows),
                "insert_position": insert_position,
                "actual_insert_row": insert_row,
                "validation_report": validation_report,
                "formula_report": formula_report,
                "formatting_preserved": preserve_formatting,
                "empty_rows_removed": removed_rows,
                "empty_rows_detected": len(empty_rows),
                "final_row_count": worksheet.max_row - 1,  # 减去标题行
                "left_aligned": True,  # 标记已应用左对齐
                "performance": {
                    "processing_time_seconds": processing_time,
                    "memory_usage_mb": memory_usage,
                    "rows_per_second": round(len(validated_rows) / processing_time, 2) if processing_time > 0 else 0
                }
            }
        }
        
        logger.info(f"[Tool] Successfully inserted {len(validated_rows)} rows, removed {removed_rows} empty rows. Processing time: {processing_time}s")
        return result
        
    except FileNotFoundError:
        error_msg = f"文件未找到: {file_path}"
        logger.error(f"[Error] {error_msg}")
        return {
            "success": False,
            "error": error_msg,
            "data": None
        }
    except PermissionError:
        error_msg = f"没有权限访问文件: {file_path}"
        logger.error(f"[Error] {error_msg}")
        return {
            "success": False,
            "error": error_msg,
            "data": None
        }
    except Exception as e:
        error_msg = f"插入Excel行时发生未知错误: {str(e)}"
        logger.error(f"[Error] {error_msg}")
        logger.error(f"[Error] Traceback: {str(e)}")
        return {
            "success": False,
            "error": error_msg,
            "data": None
        }
