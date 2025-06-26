"""
Delete Excel Row Tool - Excel行删除工具
提供根据行号删除Excel文件中指定行的功能
"""

import logging
import os
from typing import Annotated, Optional, Dict, Any
from pydantic import Field
import openpyxl

# 配置日志
logger = logging.getLogger("excel-mcp-server")


def delete_excel_row(
    file_path: Annotated[str, Field(
        description="Excel 文件的完整路径（支持 .xlsx 和 .xls 格式）"
    )],
    row_number: Annotated[int, Field(
        description="要删除的行号（从2开始计数，第1行是标题行不可删除）"
    )],
    sheet_name: Annotated[Optional[str], Field(
        default=None,
        description="要操作的工作表名称。如果不指定，将操作第一个工作表"
    )] = None
) -> Dict[str, Any]:
    """根据行号删除Excel文件中的指定行"""
    
    logger.info(f"[Tool] delete_excel_row called with file_path: {file_path}, row_number: {row_number}")
    
    try:
        # 1. 验证文件路径
        if not os.path.exists(file_path):
            error_msg = f"文件不存在: {file_path}"
            logger.error(f"[Error] {error_msg}")
            return {
                "success": False,
                "error": error_msg,
                "data": None
            }
        
        # 2. 验证文件扩展名
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext not in ['.xlsx', '.xls']:
            error_msg = f"不支持的文件格式: {file_ext}。仅支持 .xlsx 和 .xls 文件"
            logger.error(f"[Error] {error_msg}")
            return {
                "success": False,
                "error": error_msg,
                "data": None
            }
        
        # 3. 验证行号
        if row_number < 2:
            error_msg = f"行号必须大于等于2（第1行是标题行，不可删除），当前值: {row_number}"
            logger.error(f"[Error] {error_msg}")
            return {
                "success": False,
                "error": error_msg,
                "data": None
            }
        
        logger.info(f"[API] Loading Excel file: {file_path}")
        
        # 4. 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        
        # 5. 确定目标工作表
        if sheet_name and sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            target_sheet = sheet_name
        elif sheet_name:
            error_msg = f"工作表 '{sheet_name}' 不存在。可用工作表: {workbook.sheetnames}"
            logger.error(f"[Error] {error_msg}")
            return {
                "success": False,
                "error": error_msg,
                "data": None
            }
        else:
            worksheet = workbook.active
            target_sheet = worksheet.title
        
        logger.info(f"[API] Working with sheet: {target_sheet}")
        
        # 6. 检查行号是否在有效范围内
        max_row = worksheet.max_row
        if row_number > max_row:
            error_msg = f"行号 {row_number} 超出数据范围。当前工作表最大行数: {max_row}"
            logger.error(f"[Error] {error_msg}")
            return {
                "success": False,
                "error": error_msg,
                "data": None
            }
        
        # 7. 检查是否有数据可删除
        if max_row <= 1:
            error_msg = "工作表中没有数据行可以删除（只有标题行或工作表为空）"
            logger.error(f"[Error] {error_msg}")
            return {
                "success": False,
                "error": error_msg,
                "data": None
            }
        
        logger.info(f"[API] Deleting row {row_number} from sheet '{target_sheet}'")
        
        # 8. 执行删除操作
        worksheet.delete_rows(row_number)
        
        # 9. 保存文件
        workbook.save(file_path)
        logger.info(f"[API] File saved: {file_path}")
        
        # 10. 计算剩余行数
        remaining_rows = max_row - 1  # 减去删除的一行
        data_rows = remaining_rows - 1  # 减去标题行
        
        # 11. 构建返回结果
        result = {
            "success": True,
            "error": None,
            "data": {
                "file_path": file_path,
                "sheet_name": target_sheet,
                "deleted_row_number": row_number,
                "remaining_rows_count": remaining_rows,
                "data_rows_count": data_rows
            }
        }
        
        logger.info(f"[Tool] Successfully deleted row {row_number}. Remaining rows: {remaining_rows}")
        return result
        
    except FileNotFoundError:
        error_msg = f"文件未找到: {file_path}"
        logger.error(f"[Error] {error_msg}")
        return {
            "success": False,
            "error": error_msg,
            "data": None
        }
    except PermissionError:
        error_msg = f"没有权限访问文件: {file_path}"
        logger.error(f"[Error] {error_msg}")
        return {
            "success": False,
            "error": error_msg,
            "data": None
        }
    except Exception as e:
        error_msg = f"删除Excel行时发生未知错误: {str(e)}"
        logger.error(f"[Error] {error_msg}")
        return {
            "success": False,
            "error": error_msg,
            "data": None
        }
